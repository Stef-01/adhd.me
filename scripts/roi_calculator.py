#!/usr/bin/env python3
"""W21: generate the Meherr ROI calculator as a living .xlsx.

The workbook mirrors src/economics/roi.ts exactly: an Assumptions sheet of
editable inputs (blue = practice levers, yellow = fixed model assumptions) and a
Results sheet whose cells are real Excel formulas referencing the assumptions, so
a practice can change an input and watch every number recalculate. Revenue is
counted on incremental visits only — never the naive generated-booking count.

Run:  python3 scripts/roi_calculator.py
then: python3 <xlsx-skill>/scripts/recalc.py reports/roi-calculator.xlsx
Keep the defaults in sync with BRIEF_ASSUMPTIONS in src/economics/roi.ts.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BLUE = Font(name="Arial", color="0000FF")          # practice levers
BLACK = Font(name="Arial", color="000000")         # formulas / labels
BOLD = Font(name="Arial", bold=True)
TITLE = Font(name="Arial", bold=True, size=14)
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")  # fixed model assumptions
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
AUD = "$#,##0"
AUD1 = "$#,##0.0"
PCT = "0%"
NUM1 = "0.0"

wb = Workbook()

# ---- Assumptions -----------------------------------------------------------
a = wb.active
a.title = "Assumptions"
a["A1"] = "Meherr ROI calculator — assumptions"
a["A1"].font = TITLE
a["A2"] = "Blue cells are practice levers you can change. Yellow cells are the model's fixed assumptions."
a["A2"].font = Font(name="Arial", italic=True, size=9)

# label, cell, value, fmt, lever?(blue+editable) else fixed(yellow), note
rows = [
    ("Participating GPs", "B4", 10, "0", True, "Practice lever"),
    ("Bookable slots per GP per week", "B5", 120, "0", True, "24 slots/day x 5 days (W3)"),
    ("Share of capacity unfilled", "B6", 0.08, PCT, False, "W3 synthetic calibration"),
    ("Share of open slots Meherr fills", "B7", 0.25, PCT, False, "W12 sim response rate"),
    ("Did-not-attend rate on generated bookings", "B8", 0.05, PCT, False, "W12 sim"),
    ("Share of attended visits that are incremental", "B9", 0.60, PCT, False, "holdout-adjusted; not displaced"),
    ("Practice billing per attended visit (AUD)", "B10", 80, AUD, True, "W20 default; practice-set"),
    ("Meherr subscription (AUD/month)", "B11", 990, AUD, True, "pricing assumption — finalised W47"),
]
a["A3"], a["B3"] = "Input", "Value"
for c in ("A3", "B3", "C3"):
    a[c].fill = HEAD_FILL
    a[c].font = HEAD_FONT
a["C3"] = "Source / note"
for label, cell, value, fmt, lever, note in rows:
    r = cell[1:]
    a[f"A{r}"] = label
    a[f"A{r}"].font = BLACK
    a[cell] = value
    a[cell].number_format = fmt
    a[cell].font = BLUE if lever else BLACK
    a[cell].fill = YELLOW_FILL if not lever else PatternFill()
    a[f"C{r}"] = note
    a[f"C{r}"].font = Font(name="Arial", size=9, color="6B7280")
for col, width in (("A", 42), ("B", 14), ("C", 34)):
    a.column_dimensions[col].width = width

# ---- Results (formulas only) ----------------------------------------------
r = wb.create_sheet("Results")
r["A1"] = "Results"
r["A1"].font = TITLE
r["A3"], r["B3"] = "Measure", "Value"
for c in ("A3", "B3"):
    r[c].fill = HEAD_FILL
    r[c].font = HEAD_FONT

# Every value is a formula referencing Assumptions — the sheet stays live.
A = "Assumptions"
results = [
    ("Unfilled slots per week", f"={A}!B4*{A}!B5*{A}!B6", NUM1),
    ("Bookings generated per week", "=B4*Assumptions!B7", NUM1),
    ("Attended per week (after DNA)", "=B5*(1-Assumptions!B8)", NUM1),
    ("Incremental attended per week", "=B6*Assumptions!B9", NUM1),
    ("Incremental revenue per week", "=B7*Assumptions!B10", AUD),
    ("Incremental visits per year", "=B7*52", NUM1),
    ("Incremental revenue per year", "=B8*52", AUD),
    ("Meherr cost per year", "=Assumptions!B11*12", AUD),
    ("Net annual benefit", "=B10-B11", AUD),
    ("Return on cost (x)", "=IF(B11=0,0,B10/B11)", "0.0\\x"),
]
row = 4
for label, formula, fmt in results:
    r[f"A{row}"] = label
    r[f"A{row}"].font = BLACK
    r[f"B{row}"] = formula
    r[f"B{row}"].number_format = fmt
    r[f"B{row}"].font = BOLD if label in ("Net annual benefit", "Return on cost (x)") else BLACK
    for c in (f"A{row}", f"B{row}"):
        r[c].border = BORDER
    row += 1
r["A15"] = "Revenue is counted on incremental visits only — never the naive generated-booking count."
r["A15"].font = Font(name="Arial", italic=True, size=9, color="6B7280")
for col, width in (("A", 34), ("B", 16)):
    r.column_dimensions[col].width = width
r["A2"] = "Change any blue input on the Assumptions sheet and every figure here recalculates."
r["A2"].font = Font(name="Arial", italic=True, size=9)
r.sheet_view.showGridLines = False
a.sheet_view.showGridLines = False

# Force a full recalculation when the workbook is opened, so the living formulas
# populate in the practice's Excel even though openpyxl writes no cached values
# (LibreOffice headless recalc is unavailable in the build sandbox).
wb.calculation.fullCalcOnLoad = True

wb.save("reports/roi-calculator.xlsx")
print("wrote reports/roi-calculator.xlsx")
